"""RAG 검색 성능 평가: Hit Rate@K, MRR 측정.

하는 일:
    테스트셋(xlsx)의 각 쿼리로 pipeline.search()를 실행하고,
    정답 특허가 Top-K 안에 있는지 확인하여 성능 지표를 계산합니다.

    평가 실행 시 eval/reports/ 폴더에 텍스트 보고서가 자동 저장됩니다.

관계:
    - pipeline.py의 search()를 호출하여 검색 수행
    - eval/interactive_test.py의 --eval 모드에서 사용
"""
import openpyxl
from datetime import datetime
from pathlib import Path

from .. import config
from ..pipeline import search

REPORT_DIR = Path(__file__).parent / "reports"


def load_test_dataset(path: str | Path = None) -> list[dict]:
    """test_dataset xlsx 로드."""
    path = Path(path) if path else config.PROJECT_DIR / "test_dataset_18_40개.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(headers, row))
        rows.append(entry)

    wb.close()
    return rows


def evaluate(
    test_data: list[dict] = None,
    top_ks: list[int] = None,
    verbose: bool = True,
    **search_kwargs,
) -> dict:
    """Hit Rate@K, MRR 평가."""
    if test_data is None:
        test_data = load_test_dataset()

    top_ks = top_ks or [1, 3, 5, 10]
    max_k = max(top_ks)

    hits = {k: 0 for k in top_ks}
    rr_sum = 0.0
    details = []
    total = len(test_data)

    for i, row in enumerate(test_data):
        query = row.get("user_query", "")
        answer_regit = str(row.get("regit_num", "")).strip()
        answer_apply = str(row.get("apply_num", "")).strip()

        if not query:
            continue

        results = search(query, top_k=max_k, **search_kwargs)

        rank = None
        for r_idx, result in enumerate(results):
            meta = result.get("metadata", {})
            result_regit = str(meta.get("regit_num", "")).strip()
            result_apply = str(result.get("patent_id", "")).strip()

            if (answer_regit and answer_regit in result_regit) or \
               (answer_apply and answer_apply == result_apply):
                rank = r_idx + 1
                break

        for k in top_ks:
            if rank is not None and rank <= k:
                hits[k] += 1

        if rank is not None:
            rr_sum += 1.0 / rank

        top_results = []
        for r in results[:5]:
            meta = r.get("metadata", {})
            top_results.append({
                "patent_id": r.get("patent_id", "?"),
                "score": r.get("score", 0),
                "title": meta.get("invention_title", "")[:40],
                "matched_claim": r.get("matched_claim_num", "?"),
            })

        detail = {
            "idx": i,
            "query": query,
            "answer_apply": answer_apply,
            "answer_regit": answer_regit,
            "rank": rank,
            "top5": top_results,
        }
        details.append(detail)

        if verbose:
            status = f"HIT rank={rank}" if rank else "MISS"
            print(f"\n  [{i+1}/{total}] {status}")
            print(f"  쿼리: {query[:80]}")
            print(f"  정답: apply={answer_apply}, regit={answer_regit}")
            print(f"  {'순위':<4} {'특허ID':<20} {'점수':<10} {'항':<4} {'제목'}")
            print(f"  {'-'*65}")
            for j, tr in enumerate(top_results):
                marker = " <-- 정답" if rank and j + 1 == rank else ""
                print(f"  {j+1:<4} {tr['patent_id']:<20} {tr['score']:<10.6f} {tr['matched_claim']:<4} {tr['title']}{marker}")

    result = {
        "total": total,
        "mrr": rr_sum / total if total > 0 else 0,
    }
    for k in top_ks:
        result[f"hit_rate@{k}"] = hits[k] / total if total > 0 else 0

    result["details"] = details

    if verbose:
        print(f"\n{'='*50}")
        print(f"총 {total}개 쿼리")
        for k in top_ks:
            print(f"  Hit Rate@{k}: {result[f'hit_rate@{k}']:.1%} ({hits[k]}/{total})")
        print(f"  MRR: {result['mrr']:.4f}")

    report_path = _save_report(result, top_ks, hits, total)
    if report_path:
        print(f"\n  보고서 저장: {report_path}")

    return result


def _save_report(result: dict, top_ks: list[int], hits: dict, total: int) -> Path | None:
    """평가 결과를 텍스트 보고서로 저장."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = REPORT_DIR / f"eval_{timestamp}.txt"

    lines = []
    lines.append("=" * 70)
    lines.append(f"RAG 평가 보고서  ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})")
    lines.append("=" * 70)
    lines.append("")
    lines.append("[요약]")
    lines.append(f"  총 쿼리: {total}개")
    for k in top_ks:
        lines.append(f"  Hit Rate@{k}: {result[f'hit_rate@{k}']:.1%} ({hits[k]}/{total})")
    lines.append(f"  MRR: {result['mrr']:.4f}")
    lines.append("")
    lines.append("[상세 결과]")
    for d in result["details"]:
        idx = d["idx"] + 1
        rank = d["rank"]
        status = f"HIT rank={rank}" if rank else "MISS"
        lines.append("")
        lines.append(f"--- [{idx}/{total}] {status} ---")
        lines.append(f"  쿼리: {d['query']}")
        lines.append(f"  정답: apply={d.get('answer_apply','')}, regit={d.get('answer_regit','')}")
        lines.append(f"  {'순위':<4} {'특허ID':<22} {'점수':<10} {'항':<4} {'제목'}")
        lines.append(f"  {'-'*68}")
        for j, tr in enumerate(d.get("top5", [])):
            marker = " <-- 정답" if rank and j + 1 == rank else ""
            score = tr['score'] if isinstance(tr['score'], (int, float)) else 0
            lines.append(f"  {j+1:<4} {tr['patent_id']:<22} {score:<10.6f} {str(tr.get('matched_claim','?')):<4} {tr.get('title','')}{marker}")

    lines.append("")
    lines.append("=" * 70)

    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path
