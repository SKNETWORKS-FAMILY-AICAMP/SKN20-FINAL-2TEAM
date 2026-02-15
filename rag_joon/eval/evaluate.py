"""RAG 검색 성능 평가: Hit Rate@K, MRR 측정 (label별 분석 포함).

하는 일:
    테스트셋(xlsx)의 각 쿼리로 pipeline.search()를 실행하고,
    정답 특허가 Top-K 안에 있는지 확인하여 성능 지표를 계산합니다.

    필터 조건:
        - is_correct = "O"
        - label != "비침해" (침해, 침해_전문가, 애매만 평가)

    평가 기준:
        - 침해 / 침해_전문가: 반드시 검색 결과에 포함되어야 함
        - 애매: 나오면 좋고, 안 나와도 OK (참고용)
        - Top 1 = 이상적 / Top 3 = 매우 좋음 / Top 5 = 좋음

    평가 실행 시 eval/reports/ 폴더에 텍스트 보고서가 자동 저장됩니다.

관계:
    - pipeline.py의 search()를 호출하여 검색 수행
    - eval/interactive_test.py의 --eval 모드에서 사용
"""
import time
import openpyxl
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from .. import config
from ..pipeline import search

# ══════════════════════════════════════════════════════
# 테스트 데이터셋 로드
# ══════════════════════════════════════════════════════

REPORT_DIR = Path(__file__).parent / "reports"
DEFAULT_DATASET = Path(__file__).parent / "test_dataset_merged_116.xlsx"


def load_test_dataset(path: str | Path = None) -> list[dict]:
    """test_dataset xlsx 로드 + 필터 적용.

    필터: is_correct='O' AND label != '비침해'
    """
    path = Path(path) if path else DEFAULT_DATASET
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    # 첫 번째 행에서 헤더 추출
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = dict(zip(headers, row))
        is_correct = str(entry.get("is_correct", "")).strip()
        label = str(entry.get("label", "")).strip()

        # 필터: 정답 아닌 행과 비침해 라벨 제외
        if is_correct != "O" or label == "비침해":
            skipped += 1
            continue

        # 질문 생성 불가 행 제외
        user_query = str(entry.get("user_query", "")).strip()
        if "애매 질문 생성 어려움" in user_query:
            skipped += 1
            continue

        rows.append(entry)

    wb.close()
    print(f"  데이터셋 로드: {path.name}")
    print(f"  전체 → 필터 후: {len(rows) + skipped} → {len(rows)}건 (제외: {skipped}건)")
    print(f"  label 분포: {_count_labels(rows)}")
    return rows


def _count_labels(data: list[dict]) -> str:
    """label별 건수를 "침해=30, 애매=15" 형태 문자열로 반환."""
    counts = defaultdict(int)
    for row in data:
        counts[str(row.get("label", "?")).strip()] += 1
    return ", ".join(f"{k}={v}" for k, v in sorted(counts.items()))



# ══════════════════════════════════════════════════════
# 평가 실행 (Hit Rate@K, MRR)
# ══════════════════════════════════════════════════════

def evaluate(
    test_data: list[dict] = None,
    top_ks: list[int] = None,
    verbose: bool = True,
    **search_kwargs,
) -> dict:
    """Hit Rate@K, MRR 평가 (label별 분석 포함)."""
    if test_data is None:
        test_data = load_test_dataset()

    top_ks = top_ks or [1, 3, 5, 10]
    max_k = max(top_ks)

    # 전체 통계 (Hit@K별 적중 수, MRR용 역순위 합산)
    hits = {k: 0 for k in top_ks}
    rr_sum = 0.0
    details = []
    total = len(test_data)

    # label별 통계 (침해/침해_전문가/애매 각각의 적중률 분석)
    label_hits = defaultdict(lambda: {k: 0 for k in top_ks})
    label_total = defaultdict(int)
    label_rr_sum = defaultdict(float)

    for i, row in enumerate(test_data):
        query = row.get("user_query", "")
        answer_regit = str(row.get("regit_num", "")).strip()
        answer_apply = str(row.get("apply_num", "")).strip()
        label = str(row.get("label", "")).strip()

        if not query:
            continue

        t0 = time.perf_counter()
        results = search(query, top_k=max_k, **search_kwargs)
        elapsed = time.perf_counter() - t0

        # 검색 결과에서 정답 특허의 순위 탐색
        rank = None
        for r_idx, result in enumerate(results):
            meta = result.get("metadata", {})
            result_regit = str(meta.get("regit_num", "")).strip()
            result_apply = str(result.get("patent_id", "")).strip()

            # 하이픈 제거 후 비교 (출원번호 형식 통일)
            result_apply_norm = result_apply.replace("-", "")
            answer_apply_norm = answer_apply.replace("-", "")

            # 등록번호 또는 출원번호 일치 시 정답으로 판정
            if (answer_regit and answer_regit in result_regit) or \
               (answer_apply_norm and answer_apply_norm == result_apply_norm):
                rank = r_idx + 1
                break

        # 전체 통계 업데이트
        for k in top_ks:
            if rank is not None and rank <= k:
                hits[k] += 1
                label_hits[label][k] += 1

        if rank is not None:
            rr_sum += 1.0 / rank
            label_rr_sum[label] += 1.0 / rank

        label_total[label] += 1

        top_results = []
        for r in results[:max_k]:
            meta = r.get("metadata", {})
            top_results.append({
                "patent_id": r.get("patent_id", "?"),
                "regit_num": meta.get("regit_num", "?"),
                "score": r.get("score", 0),
                "title": meta.get("invention_title", "")[:40],
                "matched_claim": r.get("matched_claim_num", "?"),
            })

        detail = {
            "idx": i,
            "query": query,
            "label": label,
            "answer_apply": answer_apply,
            "answer_regit": answer_regit,
            "rank": rank,
            "elapsed": elapsed,
            "top_results": top_results,
        }
        details.append(detail)

        if verbose:
            grade = _rank_grade(rank)
            status = f"HIT rank={rank} ({grade})" if rank else "MISS"
            print(f"\n  [{i+1}/{total}] [{label}] {status} ({elapsed:.3f}s)")
            print(f"  쿼리: {query[:80]}")
            print(f"  정답: regit={answer_regit}")
            print(f"  {'순위':<4} {'출원번호':<20} {'등록번호':<22} {'점수':<10} {'항':<4} {'제목'}")
            print(f"  {'-'*85}")
            for j, tr in enumerate(top_results):
                marker = " <-- 정답" if rank and j + 1 == rank else ""
                print(f"  {j+1:<4} {tr['patent_id']:<20} {tr['regit_num']:<22} {tr['score']:<10.6f} {tr['matched_claim']:<4} {tr['title']}{marker}")

    # 결과 집계
    result = {
        "total": total,
        "mrr": rr_sum / total if total > 0 else 0,
    }
    for k in top_ks:
        result[f"hit_rate@{k}"] = hits[k] / total if total > 0 else 0

    result["details"] = details

    # 속도 통계
    times = [d["elapsed"] for d in details]
    if times:
        result["time_avg"] = sum(times) / len(times)
        result["time_min"] = min(times)
        result["time_max"] = max(times)
        result["time_total"] = sum(times)
    else:
        result["time_avg"] = result["time_min"] = result["time_max"] = result["time_total"] = 0

    # label별 결과
    label_results = {}
    for label in sorted(label_total.keys()):
        lt = label_total[label]
        lr = {}
        for k in top_ks:
            lr[f"hit_rate@{k}"] = label_hits[label][k] / lt if lt > 0 else 0
            lr[f"hits@{k}"] = label_hits[label][k]
        lr["total"] = lt
        lr["mrr"] = label_rr_sum[label] / lt if lt > 0 else 0
        label_results[label] = lr
    result["label_results"] = label_results

    # 실제 사용된 파라미터 기록
    actual_rrf_weights = search_kwargs.get("rrf_weights", config.RRF_WEIGHTS)
    result["rrf_weights"] = actual_rrf_weights

    if verbose:
        _print_summary(result, top_ks, hits, total, label_results, label_total, label_hits, actual_rrf_weights)

    report_path = _save_report(result, top_ks, hits, total, label_results, label_total, label_hits, actual_rrf_weights)
    if report_path:
        print(f"\n  보고서 저장: {report_path}")

    return result



# ══════════════════════════════════════════════════════
# 결과 출력 + 보고서 저장
# ══════════════════════════════════════════════════════

def _rank_grade(rank: int | None) -> str:
    """순위를 사람이 읽을 수 있는 등급으로 변환."""
    if rank is None:
        return "MISS"
    if rank == 1:
        return "이상적"
    if rank <= 3:
        return "매우 좋음"
    if rank <= 5:
        return "좋음"
    if rank <= 10:
        return "참고"
    return f"Top-{rank}"


def _print_summary(result, top_ks, hits, total, label_results, label_total, label_hits, rrf_weights=None):
    """콘솔에 요약 출력."""
    rrf_weights = rrf_weights or config.RRF_WEIGHTS
    print(f"\n{'='*60}")
    print(f"  설정")
    print(f"{'─'*60}")
    print(f"  embed_model: {config.EMBED_MODEL}")
    print(f"  multi_query_mode: {config.MULTI_QUERY_MODE}")
    print(f"  rrf_weights: {rrf_weights} (dense, sparse)")
    print(f"  dense_top_k: {config.DENSE_TOP_K} / bm25_top_k: {config.BM25_TOP_K}")
    print(f"  estoppel: {config.ESTOPPEL_ENABLED} / registered_only: {config.REGISTERED_ONLY}")
    print(f"\n{'='*60}")
    print(f"  전체 결과 ({total}건)")
    print(f"{'='*60}")
    for k in top_ks:
        grade = {1: "이상적", 3: "매우 좋음", 5: "좋음", 10: "참고"}.get(k, "")
        print(f"  Hit Rate@{k} ({grade}): {result[f'hit_rate@{k}']:.1%} ({hits[k]}/{total})")
    print(f"  MRR: {result['mrr']:.4f}")

    print(f"\n{'─'*60}")
    print(f"  쿼리 응답 속도")
    print(f"{'─'*60}")
    print(f"  평균: {result['time_avg']:.3f}s")
    print(f"  최소: {result['time_min']:.3f}s")
    print(f"  최대: {result['time_max']:.3f}s")
    print(f"  합계: {result['time_total']:.1f}s")

    print(f"\n{'─'*60}")
    print(f"  label별 결과")
    print(f"{'─'*60}")

    must_hit_labels = ["침해", "침해_전문가"]
    optional_label = "애매"

    for label in sorted(label_total.keys()):
        lt = label_total[label]
        requirement = "필수" if label in must_hit_labels else "참고"
        print(f"\n  [{label}] ({lt}건) - {requirement}")
        for k in top_ks:
            h = label_hits[label][k]
            rate = h / lt if lt > 0 else 0
            print(f"    Hit Rate@{k}: {rate:.1%} ({h}/{lt})")
        mrr = label_results[label]["mrr"]
        print(f"    MRR: {mrr:.4f}")

    # MISS 목록
    miss_details = [d for d in result["details"] if d["rank"] is None]
    if miss_details:
        print(f"\n{'─'*60}")
        print(f"  MISS 목록 ({len(miss_details)}건)")
        print(f"{'─'*60}")
        for d in miss_details:
            print(f"  [{d['label']}] {d['query'][:60]}  (정답: {d['answer_regit']})")

    print(f"\n{'='*60}")


def _save_report(result, top_ks, hits, total, label_results, label_total, label_hits, rrf_weights=None) -> Path | None:
    """평가 결과를 텍스트 보고서로 저장."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = REPORT_DIR / f"eval_{timestamp}.txt"

    lines = []
    lines.append("=" * 70)
    lines.append(f"RAG 평가 보고서  ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})")
    lines.append("=" * 70)
    lines.append("")

    # 설정값
    rrf_weights = rrf_weights or config.RRF_WEIGHTS
    lines.append("[설정]")
    lines.append(f"  embed_model: {config.EMBED_MODEL}")
    lines.append(f"  dense_top_k: {config.DENSE_TOP_K}")
    lines.append(f"  bm25_top_k: {config.BM25_TOP_K}")
    lines.append(f"  rrf_weights: {rrf_weights} (dense, sparse)")
    lines.append(f"  rrf_k: {config.RRF_K}")
    lines.append(f"  final_top_k: {config.FINAL_TOP_K}")
    lines.append(f"  multi_query_mode: {config.MULTI_QUERY_MODE}")
    lines.append(f"  estoppel_enabled: {config.ESTOPPEL_ENABLED}")
    lines.append(f"  registered_only: {config.REGISTERED_ONLY}")
    lines.append("")

    # 전체 요약
    lines.append("[전체 요약]")
    lines.append(f"  총 쿼리: {total}개")
    lines.append(f"  필터: is_correct=O, label!=비침해")
    for k in top_ks:
        grade = {1: "이상적", 3: "매우 좋음", 5: "좋음", 10: "참고"}.get(k, "")
        lines.append(f"  Hit Rate@{k} ({grade}): {result[f'hit_rate@{k}']:.1%} ({hits[k]}/{total})")
    lines.append(f"  MRR: {result['mrr']:.4f}")
    lines.append("")

    # 속도 통계
    lines.append("[쿼리 응답 속도]")
    lines.append(f"  평균: {result['time_avg']:.3f}s")
    lines.append(f"  최소: {result['time_min']:.3f}s")
    lines.append(f"  최대: {result['time_max']:.3f}s")
    lines.append(f"  합계: {result['time_total']:.1f}s")
    lines.append("")

    # label별 요약
    lines.append("[label별 요약]")
    must_hit_labels = ["침해", "침해_전문가"]
    for label in sorted(label_total.keys()):
        lt = label_total[label]
        requirement = "필수" if label in must_hit_labels else "참고"
        lines.append(f"  [{label}] ({lt}건) - {requirement}")
        for k in top_ks:
            h = label_hits[label][k]
            rate = h / lt if lt > 0 else 0
            lines.append(f"    Hit Rate@{k}: {rate:.1%} ({h}/{lt})")
        lines.append(f"    MRR: {label_results[label]['mrr']:.4f}")
    lines.append("")

    # MISS 목록
    miss_details = [d for d in result["details"] if d["rank"] is None]
    if miss_details:
        lines.append(f"[MISS 목록] ({len(miss_details)}건)")
        for d in miss_details:
            lines.append(f"  [{d['label']}] {d['query'][:80]}")
            lines.append(f"    정답: regit={d['answer_regit']}, apply={d['answer_apply']}")
        lines.append("")

    # 상세 결과
    lines.append("[상세 결과]")
    for d in result["details"]:
        idx = d["idx"] + 1
        rank = d["rank"]
        grade = _rank_grade(rank)
        status = f"HIT rank={rank} ({grade})" if rank else "MISS"
        lines.append("")
        lines.append(f"--- [{idx}/{total}] [{d['label']}] {status} ({d.get('elapsed', 0):.3f}s) ---")
        lines.append(f"  쿼리: {d['query']}")
        lines.append(f"  정답: apply={d.get('answer_apply','')}, regit={d.get('answer_regit','')}")
        lines.append(f"  {'순위':<4} {'출원번호':<22} {'등록번호':<22} {'점수':<10} {'항':<4} {'제목'}")
        lines.append(f"  {'-'*88}")
        for j, tr in enumerate(d.get("top_results", [])):
            marker = " <-- 정답" if rank and j + 1 == rank else ""
            score = tr['score'] if isinstance(tr['score'], (int, float)) else 0
            lines.append(f"  {j+1:<4} {tr['patent_id']:<22} {tr.get('regit_num','?'):<22} {score:<10.6f} {str(tr.get('matched_claim','?')):<4} {tr.get('title','')}{marker}")

    lines.append("")
    lines.append("=" * 70)

    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path
